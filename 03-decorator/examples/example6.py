import unittest
import nose
import os
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import inspect


def load_html_file(path):
    """Call test with a html file of the same name.

    Args:
        path: Relative path where the html file is located."""

    def test_decorator(fn):
        base_path = os.path.join(os.path.dirname(__file__), path)
        file_name = fn.__name__ + ".html"
        file_path = os.path.join(base_path, file_name)

        html_f = open(file_path, "r")

        def test_decorated(self, *args, **kwargs):
            fn(self, html_f=html_f, *args, **kwargs)

        test_decorated.__name__ = fn.__name__
        return test_decorated
    return test_decorator


def load_excel_file(path):
    """Call test with a excel file of the same name.

    Args:
        path: Relative path where the html file is located."""

    def test_decorator(fn):
        base_path = os.path.join(os.path.dirname(__file__), path)
        file_name = fn.__name__.replace("html", "excel") + ".xlsx"
        file_path = os.path.join(base_path, file_name)

        wb = load_workbook(file_path)

        def test_decorated(self, *args, **kwargs):
            fn(self, wb=wb, *args, **kwargs)

        test_decorated.__name__ = fn.__name__
        return test_decorated
    return test_decorator


class SomeScrapingTest(unittest.TestCase):

    @load_excel_file("mock_excel_files")
    @load_html_file("mock_html_files")
    def test_mock_html_1(self, html_f, wb):

        scrape_this = "I am mock html number 1, scrape me!"

        bs = BeautifulSoup(html_f)
        test_scraped = bs.find("p").get_text()

        ws = wb.active
        num_test = int(inspect.stack()[0][3].split("_")[-1])
        check_this = ws.cell(row=num_test, column=num_test).value

        self.assertEqual(test_scraped, scrape_this)
        self.assertEqual(check_this, scrape_this)

    @load_excel_file("mock_excel_files")
    @load_html_file("mock_html_files")
    def test_mock_html_2(self, html_f, wb):

        scrape_this = "I am mock html number 2, scrape me!"

        bs = BeautifulSoup(html_f)
        test_scraped = bs.find("p").get_text()

        ws = wb.active
        num_test = int(inspect.stack()[0][3].split("_")[-1])
        check_this = ws.cell(row=num_test, column=num_test).value

        self.assertEqual(test_scraped, scrape_this)
        self.assertEqual(check_this, scrape_this)

    @load_html_file("mock_html_files")
    def test_mock_html_3(self, html_f, wb):

        scrape_this = "I am mock html number 3, scrape me!"

        bs = BeautifulSoup(html_f)
        test_scraped = bs.find("p").get_text()

        self.assertEqual(test_scraped, scrape_this)

    @load_html_file("mock_html_files")
    def test_mock_html_4(self, html_f):

        scrape_this = "I am mock html number 4, scrape me!"

        bs = BeautifulSoup(html_f)
        test_scraped = bs.find("p").get_text()

        self.assertEqual(test_scraped, scrape_this)


if __name__ == '__main__':
    nose.run(defaultTest=__name__)
